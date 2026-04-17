#!/usr/bin/env python3
"""
从Excel文件生成 report.json
用法: python generate_report_json.py [blood_excel] [ldh_cart_excel] [output_json]
默认:
  blood_excel = 血常规报告汇总.xlsx
  ldh_cart_excel = LDH_Cart.xlsx
  output_json = src/data/report.json
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误: 请先安装 openpyxl")
    print("pip install openpyxl")
    sys.exit(1)


def read_blood_data(excel_path):
    """读取血常规数据"""
    wb = load_workbook(excel_path, data_only=True)

    # 尝试找到包含数据的sheet
    ws = None
    for sheet_name in wb.sheetnames:
        if '汇总' in sheet_name or '数据' in sheet_name:
            ws = wb[sheet_name]
            break
    if ws is None:
        # 使用第二个sheet（通常包含图表数据）
        ws = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb.active

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and any(row[1:]):  # 确保有日期和至少一个数据值
            try:
                record = {
                    'date': str(row[0])[:10] if isinstance(row[0], str) else str(row[0]),
                    'wbc': float(row[1]) if row[1] is not None else 0,
                    'ne': float(row[2]) if row[2] is not None else 0,
                    'ly': float(row[3]) if row[3] is not None else 0,
                    'plt': float(row[4]) if row[4] is not None else 0,
                    'hgb': float(row[5]) if row[5] is not None else 0
                }
                records.append(record)
            except (ValueError, TypeError):
                continue

    return records


def read_ldh_cart_data(excel_path):
    """读取LDH和Cart细胞数据"""
    wb = load_workbook(excel_path, data_only=True)

    ldh_records = []
    cart_dates = []
    cart_data = {
        'tPct': [], 'cd4Pct': [], 'cd8Pct': [],
        'tNum': [], 'cd4Num': [], 'cd8Num': [],
        'carTPct': [], 'cd4CarTPct': [], 'cd8CarTPct': [],
        'carTNum': [], 'cd4CarTNum': [], 'cd8CarTNum': []
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 读取LDH数据 (查找包含LDH的sheet或列)
        if 'LDH' in sheet_name.upper() or '乳酸' in sheet_name:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    try:
                        date_val = str(row[0])[:10] if isinstance(row[0], str) else str(row[0])
                        value = float(row[1]) if row[1] is not None else 0
                        ldh_records.append({'date': date_val, 'value': value})
                    except (ValueError, TypeError):
                        continue

        # 读取Cart数据 (查找包含CD3、CarT的sheet或列)
        if 'CAR' in sheet_name.upper() or 'CART' in sheet_name or 'T细胞' in sheet_name:
            # 第一行通常是日期
            headers = [str(c)[:10] if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            # 查找日期列
            date_indices = []
            for i, h in enumerate(headers):
                if any(x in h.upper() for x in ['2026', '2025', '日期', 'DATE']):
                    date_indices.append(i)

            if date_indices:
                # 使用第一个日期作为参考，其余是数据列
                first_date_idx = date_indices[0]
                if len(date_indices) > 1:
                    cart_dates = [headers[i] for i in date_indices[1:] if headers[i]]

                # 读取数据行
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # 第一列是指标名称
                        indicator = str(row[0]).strip()
                        key_map = {
                            'T%': 'tPct', 'CD4+%': 'cd4Pct', 'CD8+%': 'cd8Pct',
                            'T#': 'tNum', 'CD4+#': 'cd4Num', 'CD8+#': 'cd8Num',
                            'CAR-T%': 'carTPct', 'CD4+CAR-T%': 'cd4CarTPct', 'CD8+CAR-T%': 'cd8CarTPct',
                            'CAR-T#': 'carTNum', 'CD4+CAR-T#': 'cd4CarTNum', 'CD8+CAR-T#': 'cd8CarTNum'
                        }

                        for k, v in key_map.items():
                            if k.upper() in indicator.upper():
                                if len(date_indices) > 1:
                                    cart_data[v] = [float(row[i]) if row[i] is not None else 0 for i in date_indices[1:]]
                                break

    return ldh_records, cart_dates, cart_data


def generate_report_json(blood_excel, ldh_cart_excel, output_path):
    """生成标准化的report.json"""
    report = {
        'meta': {
            'generatedAt': datetime.now().isoformat(),
            'sourceFiles': [blood_excel, ldh_cart_excel]
        },
        'blood': {
            'ranges': {
                'wbc': {'min': 3.5, 'max': 9.5, 'unit': '×10⁹/L'},
                'ne': {'min': 1.8, 'max': 6.3, 'unit': '×10⁹/L'},
                'ly': {'min': 1.1, 'max': 3.2, 'unit': '×10⁹/L'},
                'plt': {'min': 125, 'max': 350, 'unit': '×10⁹/L'},
                'hgb': {'min': 115, 'max': 150, 'unit': 'g/L'}
            },
            'records': []
        },
        'ldh': {
            'ranges': {'min': 120, 'max': 250, 'unit': 'U/L'},
            'records': []
        },
        'cart': {
            'dates': [],
            'records': {
                'tPct': {'label': '总T细胞(T%)', 'unit': '%', 'reference': '51~85', 'values': []},
                'cd4Pct': {'label': '辅助T细胞(CD4+%)', 'unit': '%', 'reference': '33~58', 'values': []},
                'cd8Pct': {'label': '杀伤T细胞(CD8+%)', 'unit': '%', 'reference': '9~39', 'values': []},
                'tNum': {'label': '总T细胞(T#)', 'unit': '个/μL', 'reference': '723~2737', 'values': []},
                'cd4Num': {'label': '辅助T细胞(CD4+##)', 'unit': '个/μL', 'reference': '404~1612', 'values': []},
                'cd8Num': {'label': '杀伤T细胞(CD8+##)', 'unit': '个/μL', 'reference': '220~1129', 'values': []},
                'carTPct': {'label': 'CarT-T(CD3+CarT%)', 'unit': '%', 'reference': '0~0', 'values': []},
                'cd4CarTPct': {'label': 'CD4+Car-T(%)', 'unit': '%', 'reference': '0~0', 'values': []},
                'cd8CarTPct': {'label': 'CD8+Car-T(%)', 'unit': '%', 'reference': '0~0', 'values': []},
                'carTNum': {'label': 'CarT-T#', 'unit': '个/μL', 'reference': '0~0', 'values': []},
                'cd4CarTNum': {'label': 'CD4+Car-T#', 'unit': '个/μL', 'reference': '0~0', 'values': []},
                'cd8CarTNum': {'label': 'CD8+Car-T#', 'unit': '个/μL', 'reference': '0~0', 'values': []}
            }
        }
    }

    # 读取血常规数据
    if Path(blood_excel).exists():
        report['blood']['records'] = read_blood_data(blood_excel)
        print(f"✓ 已读取血常规数据: {len(report['blood']['records'])} 条记录")
    else:
        print(f"⚠ 血常规文件不存在: {blood_excel}")

    # 读取LDH和Cart数据
    if Path(ldh_cart_excel).exists():
        ldh_records, cart_dates, cart_data = read_ldh_cart_data(ldh_cart_excel)
        report['ldh']['records'] = ldh_records
        report['cart']['dates'] = cart_dates
        report['cart']['records'] = cart_data
        print(f"✓ 已读取LDH数据: {len(ldh_records)} 条记录")
        print(f"✓ 已读取Cart数据: {len(cart_dates)} 个日期")
    else:
        print(f"⚠ LDH/Cart文件不存在: {ldh_cart_excel}")

    # 确保输出目录存在
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # 写入JSON文件
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"✓ report.json已生成: {output_path}")
    return report


if __name__ == '__main__':
    # 默认参数
    blood_excel = '血常规报告汇总.xlsx'
    ldh_cart_excel = 'LDH_Cart.xlsx'
    output_json = 'src/data/report.json'

    # 从命令行参数覆盖
    if len(sys.argv) > 1:
        blood_excel = sys.argv[1]
    if len(sys.argv) > 2:
        ldh_cart_excel = sys.argv[2]
    if len(sys.argv) > 3:
        output_json = sys.argv[3]

    generate_report_json(blood_excel, ldh_cart_excel, output_json)
